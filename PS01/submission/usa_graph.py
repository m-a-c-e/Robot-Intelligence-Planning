'''
Author: Manan Patel
'''

# This script will be used to set up an undirected graph
# between the states of USA using networkxxx library

import networkx as nx
from openpyxl import Workbook, load_workbook

def draw_graph(graph, nodes_dict):
    '''
    description: creates an undirected graph based on the input
                 dictionary
    graph:       instance of an undirected graph
    nodes_dict:  dictionary of parent(key) and child node list(value)
    returns:     NA
    '''
    tuple_list = []
    for node_key in nodes_dict:
        for node in nodes_dict[node_key]:
            t = (node_key, node)
            tuple_list.append(t)

    graph.add_edges_from(tuple_list)


def create_dict(filename):
    '''
    description: creates a dictionary based on the input excel sheet.
                 first column is considered key and the rest on the same
                 row make up a list as its value
    input:  excel file name
    returns: dictionary 
    '''
    wb = load_workbook(filename=filename)
    sh = wb.active
    row_ct = sh.max_row
    col_ct = sh.max_column
    result = {}
    for i in range(2, row_ct):
        dict_key = sh.cell(row=i, column=1).value
        value_list = []
        for j in range(2, col_ct):
            if sh.cell(row=i, column=j).value == None:
                break
            value_list.append(sh.cell(row=i, column=j).value)
        result.update({dict_key: value_list})
    return result


def load_graph():
    '''
    '''
    G = nx.DiGraph()                                # defines the graph structure
    G = G.to_undirected()
    state_neighbors = create_dict("Book1.xlsx")     # create the required dictionary
    draw_graph(G, state_neighbors)                  # add nodes to the dictionary
    return G
    # uncomment to see the graph
    '''
    pos = nx.spring_layout(G)
    nx.draw_networkx_nodes(G, pos, node_size=100)
    nx.draw_networkx_edges(G, pos, edgelist=G.edges(), edge_color='black')
    nx.draw_networkx_labels(G, pos, font_size=5)
    plt.show()
    '''